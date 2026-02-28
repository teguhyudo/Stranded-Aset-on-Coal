"""
02_clean_bank_data.py -- Merge and clean bank data from FS and LSEG
====================================================================
Merges bank_ratios_from_fs.csv (gold standard) with LSEG profiles and
cost of capital data. Reads the authoritative rounding/currency metadata
from each source Excel file (sheet 1000000, rows 29 and 31) to determine
the IDR scale, then converts all monetary values to USD millions.

Output: output/clean_bank_panel.csv
"""

import sys
import importlib
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
setup = importlib.import_module('00_setup')

import numpy as np
import pandas as pd
import openpyxl

BANK_FS = setup.BANK_FS_DIR
BANK_LSEG = setup.BANK_LSEG_OUTPUT
OUT = setup.OUTPUT_DIR
FX = setup.FX_IDR_USD  # 15800 IDR per USD


# ============================================================
# SCALE DETECTION FROM SOURCE EXCEL METADATA
# ============================================================
# Each bank's annual report Excel file contains a metadata sheet
# named "1000000" with:
#   - Row 29, Col B: reporting currency (e.g., "Rupiah / IDR")
#   - Row 31, Col B: rounding convention (e.g., "Jutaan / In Million")
#
# We read this metadata directly to determine the scale divisor,
# rather than inferring it from LSEG reference data.
#
# Conversion: raw FS value / divisor = IDR millions, then / FX = USD millions.

# Monetary columns to convert
MONETARY_COLS = [
    'total_assets', 'total_assets_py', 'total_equity', 'total_equity_py',
    'gross_loans', 'loan_allowance', 'net_loans', 'total_deposits',
    'gross_loans_from_notes', 'interest_income', 'interest_expense',
    'net_interest_income', 'net_income', 'net_income_parent',
    'profit_before_tax', 'provision_expense', 'npl_amount',
    'mining_loans',
]

# Rounding string -> divisor to convert raw values to IDR millions
ROUNDING_MAP = {
    'jutaan':       1.0,   # "Jutaan / In Million" -> already IDR millions
    'ribuan':       1e3,   # "Ribuan / In Thousand" -> divide by 1e3
    'satuan penuh': 1e6,   # "Satuan Penuh / Full Amount" -> divide by 1e6
}


def read_fs_metadata(ticker, year, filename, base_dir):
    """
    Read the reporting currency and rounding from the source Excel file.

    Opens {base_dir}/{ticker}/{year}/{filename}, reads sheet "1000000":
      - Cell B29: currency (e.g., "Rupiah / IDR")
      - Cell B31: rounding (e.g., "Jutaan / In Million")

    Returns (currency_str, rounding_str) or (None, None) on failure.
    """
    filepath = base_dir / ticker / str(year) / filename
    if not filepath.exists():
        return None, None
    try:
        wb = openpyxl.load_workbook(str(filepath), read_only=True, data_only=True)
        if '1000000' not in wb.sheetnames:
            wb.close()
            return None, None
        ws = wb['1000000']
        currency = ws.cell(29, 2).value
        rounding = ws.cell(31, 2).value
        wb.close()
        return str(currency) if currency else None, str(rounding) if rounding else None
    except Exception:
        return None, None


def rounding_to_divisor(rounding_str):
    """
    Map a rounding string from the Excel metadata to a scale divisor.
    Returns the divisor to convert raw values to IDR millions.
    """
    if not rounding_str:
        return np.nan
    lower = rounding_str.lower()
    for key, divisor in ROUNDING_MAP.items():
        if key in lower:
            return divisor
    return np.nan


def harmonise_to_usd_millions(df, scale_col='_scale_divisor'):
    """
    Apply scale divisor to all monetary columns, converting from
    raw IDR (mixed scale) to USD millions.
    """
    df = df.copy()
    for col in MONETARY_COLS:
        if col in df.columns:
            usd_col = col + '_usd_m'
            # raw -> IDR millions -> USD millions
            df[usd_col] = df[col] / df[scale_col] / FX
    return df


def clean_ticker_lseg(df, col='Instrument'):
    """Strip .JK suffix from LSEG instrument column."""
    df = df.copy()
    if col in df.columns:
        df['ticker'] = df[col].str.replace('.JK', '', regex=False)
    return df


# ============================================================
# REPORTED CAR (TOTAL CAPITAL ADEQUACY RATIO) LOOKUP
# ============================================================
# Actual reported total CAR/KPMM from bank annual reports, OJK
# regulatory filings, and LSEG Tier 1 ratios (2024 fiscal year).
#
# Sources:
#   - "annual_report_2024": Bank annual report filed with OJK/IDX
#   - "ojk_report_2024": OJK risk exposure publication (Dec 2024)
#   - "quarterly_report_q2_2024": Bank quarterly financial report
#   - "lseg_tier1": LSEG Datastream Tier 1 Capital Ratio (conservative
#     lower bound for total CAR, as Tier 2 is excluded)
#
# Total CAR = (Tier 1 + Tier 2 Capital) / Risk-Weighted Assets.
# Where only Tier 1 is available, values are conservative proxies.

REPORTED_CAR = {
    # D-SIBs -- actual total CAR from annual reports
    'BMRI': (0.201,  'annual_report_2024'),
    'BBRI': (0.266,  'annual_report_2024'),
    'BBNI': (0.211,  'annual_report_2024'),
    'BNGA': (0.233,  'annual_report_2024'),
    'BBTN': (0.171,  'lseg_tier1'),
    'BRIS': (0.215,  'lseg_tier1'),
    # BBCA: Tier 1 = 28.9% from LSEG; 2025 total CAR = 30.4%
    'BBCA': (0.289,  'lseg_tier1'),
    # Large banks -- actual total CAR from annual reports
    'BNLI': (0.347,  'annual_report_2024'),
    'MEGA': (0.256,  'ojk_report_2024'),
    'NISP': (0.236,  'annual_report_2024'),
    'BDMN': (0.262,  'annual_report_2024'),
    'PNBN': (0.345,  'annual_report_2024'),
    'BBKP': (0.164,  'annual_report_2024'),
    # Medium banks
    'MAYA': (0.119,  'quarterly_report_q2_2024'),
    'ARTO': (0.323,  'lseg_tier1'),
    'BTPS': (0.473,  'lseg_tier1'),
    'BBYB': (0.312,  'lseg_tier1'),
}


def main():
    print("=" * 60)
    print("02: CLEANING BANK DATA")
    print("=" * 60)

    # ----------------------------------------------------------
    # 1. Load bank ratios from financial statements (GOLD STANDARD)
    # ----------------------------------------------------------
    print("\n[1/5] Loading bank_ratios_from_fs.csv...")
    fs = pd.read_csv(BANK_FS / "bank_ratios_from_fs.csv")
    print(f"  Raw rows: {len(fs)} | Unique banks: {fs['ticker'].nunique()}")
    print(f"  Years: {sorted(fs['year'].unique())}")

    # Take latest year per bank
    idx_latest = fs.groupby('ticker')['year'].idxmax()
    fs_latest = fs.loc[idx_latest].copy().reset_index(drop=True)
    print(f"  After taking latest year per bank: {len(fs_latest)} rows")

    # ----------------------------------------------------------
    # 2. Load LSEG reference data for scale detection
    # ----------------------------------------------------------
    print("\n[2/5] Loading LSEG company profiles for scale reference...")
    prof = pd.read_csv(BANK_LSEG / "company_profiles.csv")
    prof = clean_ticker_lseg(prof)

    # LSEG total assets column (already in USD millions from LSEG)
    lseg_ta_col = 'Total Assets, Reported'
    if lseg_ta_col not in prof.columns:
        # Try alternative names
        for c in prof.columns:
            if 'total' in c.lower() and 'asset' in c.lower():
                lseg_ta_col = c
                break

    lseg_ref = prof[['ticker', lseg_ta_col, 'Company Market Cap', 'Total Equity']].copy()
    lseg_ref = lseg_ref.rename(columns={
        lseg_ta_col: 'lseg_total_assets_usd_m',
        'Company Market Cap': 'market_cap_usd_m',
        'Total Equity': 'lseg_equity_usd_m',
    })
    # Drop rows with empty tickers (e.g., NAGA)
    lseg_ref = lseg_ref.dropna(subset=['ticker'])
    lseg_ref = lseg_ref[lseg_ref['ticker'].str.strip() != '']

    # Merge LSEG reference into FS for scale detection
    fs_latest = fs_latest.merge(
        lseg_ref[['ticker', 'lseg_total_assets_usd_m']],
        on='ticker', how='left'
    )

    # ----------------------------------------------------------
    # 3. Read rounding metadata from source Excel and harmonise
    # ----------------------------------------------------------
    print("\n[3/5] Reading rounding metadata from source Excel files...")

    # Read metadata from each bank's source Excel (sheet 1000000)
    fs_currencies = []
    fs_roundings = []
    fs_divisors = []
    for _, row in fs_latest.iterrows():
        currency, rounding = read_fs_metadata(
            row['ticker'], row['year'], row['file'], BANK_FS
        )
        fs_currencies.append(currency)
        fs_roundings.append(rounding)
        fs_divisors.append(rounding_to_divisor(rounding))

    fs_latest['fs_currency'] = fs_currencies
    fs_latest['fs_rounding'] = fs_roundings
    fs_latest['_scale_divisor'] = fs_divisors

    # Report results
    n_read = fs_latest['fs_rounding'].notna().sum()
    n_missing = fs_latest['fs_rounding'].isna().sum()
    print(f"  Metadata read successfully: {n_read}/{len(fs_latest)} banks")
    if n_missing > 0:
        missing = fs_latest.loc[fs_latest['fs_rounding'].isna(), 'ticker'].tolist()
        print(f"  WARNING: {n_missing} banks missing metadata: {missing}")

    scale_labels = {1.0: 'Jutaan (IDR millions)', 1e3: 'Ribuan (IDR thousands)',
                    1e6: 'Satuan Penuh (full IDR)'}
    for divisor, label in scale_labels.items():
        n = (fs_latest['_scale_divisor'] == divisor).sum()
        if n > 0:
            tickers = fs_latest.loc[fs_latest['_scale_divisor'] == divisor, 'ticker'].tolist()
            print(f"  {label}: {n} banks -- {', '.join(tickers[:10])}"
                  + (f" + {n-10} more" if n > 10 else ""))

    # Apply conversion: raw -> IDR millions -> USD millions
    fs_latest = harmonise_to_usd_millions(fs_latest)

    # Cross-check against LSEG reference (validation, not used for detection)
    check = fs_latest[['ticker', 'total_assets_usd_m', 'lseg_total_assets_usd_m']].dropna()
    if len(check) > 0:
        check['pct_diff'] = ((check['total_assets_usd_m'] - check['lseg_total_assets_usd_m'])
                             / check['lseg_total_assets_usd_m'] * 100)
        print(f"\n  LSEG cross-check (FS vs LSEG total assets):")
        print(f"    Mean absolute % diff: {check['pct_diff'].abs().mean():.1f}%")
        print(f"    Median absolute % diff: {check['pct_diff'].abs().median():.1f}%")
        outliers = check[check['pct_diff'].abs() > 30]
        if len(outliers) > 0:
            print(f"    WARNING: {len(outliers)} banks with >30% deviation:")
            for _, row in outliers.iterrows():
                print(f"      {row['ticker']}: FS={row['total_assets_usd_m']:.1f}, "
                      f"LSEG={row['lseg_total_assets_usd_m']:.1f}, "
                      f"diff={row['pct_diff']:.1f}%")

    # ----------------------------------------------------------
    # 4. Merge with LSEG cost of capital (WACC, beta)
    # ----------------------------------------------------------
    print("\n[4/5] Merging LSEG cost of capital and volatility...")
    coc = pd.read_csv(BANK_LSEG / "cost_of_capital.csv")
    coc = clean_ticker_lseg(coc)

    # Drop duplicates -- some tickers appear multiple times (multi-year data)
    # Take the first row per ticker (they have the same beta/WACC)
    coc_dedup = coc.drop_duplicates(subset=['ticker'], keep='first')

    coc_cols = {
        'Beta': 'beta',
        'WACC_Base': 'wacc_base',
        'WACC_Climate': 'wacc_climate',
        'Cost_of_Equity_Base': 'cost_of_equity_base',
        'Cost_of_Equity_Climate': 'cost_of_equity_climate',
        'Cost_of_Debt': 'cost_of_debt',
        'Market_Cap_USD_M': 'market_cap_usd_m_coc',
        'Tier1_Capital_Ratio': 'tier1_ratio',
        'Total_CAR': 'total_car',
    }
    coc_rename = {k: v for k, v in coc_cols.items() if k in coc_dedup.columns}
    coc_dedup = coc_dedup.rename(columns=coc_rename)
    merge_coc_cols = ['ticker'] + list(coc_rename.values())
    merge_coc_cols = [c for c in merge_coc_cols if c in coc_dedup.columns]
    fs_latest = fs_latest.merge(coc_dedup[merge_coc_cols], on='ticker', how='left')

    # Load market cap from profiles (more reliable than cost_of_capital)
    fs_latest = fs_latest.merge(
        lseg_ref[['ticker', 'market_cap_usd_m']],
        on='ticker', how='left'
    )

    # Stock volatility
    vol = pd.read_csv(BANK_LSEG / "stock_volatility.csv")
    vol = clean_ticker_lseg(vol)
    vol_rename = {
        'Annualized_Volatility': 'stock_volatility',
        'Sharpe_Proxy': 'sharpe_proxy',
    }
    vol = vol.rename(columns=vol_rename)
    vol_cols = ['ticker'] + [v for v in vol_rename.values() if v in vol.columns]
    fs_latest = fs_latest.merge(vol[vol_cols], on='ticker', how='left')
    print(f"  Merged beta, WACC, volatility for {fs_latest['beta'].notna().sum()} banks")

    # ----------------------------------------------------------
    # 5. Assemble final panel
    # ----------------------------------------------------------
    print("\n[5/5] Assembling clean bank panel...")

    # Select and rename final columns
    panel = pd.DataFrame()
    panel['ticker'] = fs_latest['ticker']
    panel['year'] = fs_latest['year']

    # Monetary values (USD millions from FS)
    for src_col, dst_col in [
        ('total_assets_usd_m', 'total_assets_usd_m'),
        ('total_equity_usd_m', 'total_equity_usd_m'),
        ('gross_loans_usd_m', 'gross_loans_usd_m'),
        ('net_loans_usd_m', 'net_loans_usd_m'),
        ('total_deposits_usd_m', 'total_deposits_usd_m'),
        ('net_income_usd_m', 'net_income_usd_m'),
        ('net_income_parent_usd_m', 'net_income_parent_usd_m'),
        ('profit_before_tax_usd_m', 'profit_before_tax_usd_m'),
        ('interest_income_usd_m', 'interest_income_usd_m'),
        ('interest_expense_usd_m', 'interest_expense_usd_m'),
        ('net_interest_income_usd_m', 'net_interest_income_usd_m'),
        ('provision_expense_usd_m', 'provision_expense_usd_m'),
        ('npl_amount_usd_m', 'npl_amount_usd_m'),
        ('mining_loans_usd_m', 'mining_loans_usd_m'),
        ('loan_allowance_usd_m', 'loan_allowance_usd_m'),
    ]:
        if src_col in fs_latest.columns:
            panel[dst_col] = fs_latest[src_col]

    # Ratios (dimensionless, directly from FS)
    for col in ['mining_share', 'roa', 'roe', 'nim', 'ldr',
                'npl_gross_calculated', 'npf_gross_reported', 'npf_net_reported']:
        if col in fs_latest.columns:
            panel[col] = fs_latest[col]

    # Rename NPL column for clarity
    if 'npl_gross_calculated' in panel.columns:
        panel = panel.rename(columns={'npl_gross_calculated': 'npl_gross'})

    # LSEG-sourced columns
    for col in ['market_cap_usd_m', 'beta', 'stock_volatility', 'sharpe_proxy',
                'wacc_base', 'wacc_climate', 'cost_of_equity_base',
                'cost_of_equity_climate', 'cost_of_debt',
                'tier1_ratio', 'total_car']:
        if col in fs_latest.columns:
            panel[col] = fs_latest[col]

    # Metadata columns
    panel['fs_currency'] = fs_latest['fs_currency'].values
    panel['fs_rounding'] = fs_latest['fs_rounding'].values
    panel['has_lseg_reference'] = fs_latest['lseg_total_assets_usd_m'].notna().astype(int)

    # Reported CAR from public sources
    panel['reported_car'] = panel['ticker'].map(
        {k: v[0] for k, v in REPORTED_CAR.items()}
    )
    panel['car_source'] = panel['ticker'].map(
        {k: v[1] for k, v in REPORTED_CAR.items()}
    )
    # For banks without reported CAR, note source as not_available
    panel.loc[panel['reported_car'].isna(), 'car_source'] = 'not_available'

    # ----------------------------------------------------------
    # DERIVED COLUMNS
    # ----------------------------------------------------------
    print("  Computing derived columns...")

    # Deposit-to-asset ratio
    panel['deposit_to_asset'] = (panel['total_deposits_usd_m']
                                 / panel['total_assets_usd_m'].replace(0, np.nan))

    # Loan-to-deposit ratio (alternative from USD values)
    panel['ldr_usd'] = (panel['gross_loans_usd_m']
                        / panel['total_deposits_usd_m'].replace(0, np.nan))

    # Equity-to-asset ratio
    panel['equity_ratio'] = (panel['total_equity_usd_m']
                             / panel['total_assets_usd_m'].replace(0, np.nan))

    # Mining exposure: mining_loans / gross_loans
    panel['mining_loan_share'] = (panel['mining_loans_usd_m']
                                  / panel['gross_loans_usd_m'].replace(0, np.nan))

    # Size category
    bins = [0, 500, 5000, 20000, np.inf]
    labels = ['Small', 'Medium', 'Large', 'DSIB']
    panel['size_category'] = pd.cut(
        panel['total_assets_usd_m'], bins=bins, labels=labels, right=True
    )

    # Sort by total assets descending
    panel = panel.sort_values('total_assets_usd_m', ascending=False).reset_index(drop=True)

    # ----------------------------------------------------------
    # SAVE
    # ----------------------------------------------------------
    outpath = OUT / "clean_bank_panel.csv"
    panel.to_csv(outpath, index=False)
    print(f"\nSaved: {outpath}")
    print(f"Shape: {panel.shape}")

    # ----------------------------------------------------------
    # SUMMARY
    # ----------------------------------------------------------
    print("\n--- KEY STATISTICS ---")
    print(f"Total banks: {len(panel)}")
    print(f"With market cap: {panel['market_cap_usd_m'].notna().sum()}")
    print(f"With beta: {panel['beta'].notna().sum()}")
    print(f"With WACC: {panel['wacc_base'].notna().sum()}")
    print(f"With mining loans: {(panel['mining_loans_usd_m'].notna() & (panel['mining_loans_usd_m'] > 0)).sum()}")
    print(f"With mining share: {(panel['mining_share'].notna() & (panel['mining_share'] > 0)).sum()}")

    print(f"\nAggregate (USD millions):")
    print(f"  Total assets:    ${panel['total_assets_usd_m'].sum():>12,.1f}M")
    print(f"  Total equity:    ${panel['total_equity_usd_m'].sum():>12,.1f}M")
    print(f"  Gross loans:     ${panel['gross_loans_usd_m'].sum():>12,.1f}M")
    print(f"  Total deposits:  ${panel['total_deposits_usd_m'].sum():>12,.1f}M")
    print(f"  Mining loans:    ${panel['mining_loans_usd_m'].sum(min_count=1):>12,.1f}M")
    print(f"  Total market cap:${panel['market_cap_usd_m'].sum():>12,.1f}M")

    print(f"\nTop 10 banks by total assets:")
    top10 = panel.nlargest(10, 'total_assets_usd_m')[
        ['ticker', 'year', 'total_assets_usd_m', 'market_cap_usd_m',
         'mining_loans_usd_m', 'mining_share']
    ]
    print(top10.to_string(index=False))

    print(f"\nSize distribution:")
    print(panel['size_category'].value_counts().to_string())

    return panel


if __name__ == "__main__":
    main()
